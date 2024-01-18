
from openpyxl import Workbook, load_workbook
wb=load_workbook('tests/Salary for december.xlsx')
ws=wb.active
max_row = ws.max_row
Standartbonus = ws['J4'].value
Achivmentbonus=ws['J5'].value
Productbonus =ws['J6'].value
Monthlybonus = ws['J7'].value
totalworktime = 0
totalbreaks = 0
for i in range(2, max_row + 1):
    worktime = ws['G' + str(i)].value
    breaks = ws['F' + str(i)].value
    if (type(worktime)!=str and type(breaks)!=str):
        totalworktime += worktime
        totalbreaks += breaks
totalbonus= Standartbonus + Achivmentbonus + Productbonus + Monthlybonus
breaksbrutto=totalbreaks * Standartbonus
worktimebrutto= totalworktime * totalbonus
brutto= breaksbrutto + worktimebrutto
socialaisnodoklis=brutto * 0.105 
ienakumanodoklis=(brutto*0.23)-(socialaisnodoklis*0.2)
salary=brutto-(socialaisnodoklis+ienakumanodoklis)
ws['J9'].value=salary
wb.save('result.xlsx')
print(salary)